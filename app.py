# Import necessary libraries

import streamlit as st
import pandas as pd
import numpy as np

from openpyxl import Workbook, load_workbook
from datetime import datetime

import torch.nn.functional as F
from prediction_utils import load_scaler, predict_diagnosis_MLP, predict_diagnosis_LGBM
from model import MLP, LGBM



# Check if 'stage' is not in session_state and initialize it
if 'stage' not in st.session_state:
    st.session_state.stage = 0

# Function to set the stage
def set_stage(stage):
    st.session_state.stage = stage

# Set page configuration and theme
st.set_page_config(
    page_title="UTI Diagnosis App",
    page_icon=":test_tube:",
    layout="centered",
    initial_sidebar_state="expanded"
)





# Load the pre-trained scaler
scaler = load_scaler(r"Streamlit\train_scaler.joblib")



# Sidebar information
st.sidebar.title("UTI Diagnostic System")

with st.sidebar.expander("See information", expanded=True):
    st.write("Urinary Tract Infections (UTIs) have a significant global impact, affecting millions and posing challenges to healthcare systems. The World Health Organization identifies UTIs as among the most common bacterial infections worldwide.")

st.sidebar.header("About the application")
with st.sidebar.expander("See information", expanded=True):
    st.write("Empowering healthcare providers with a streamlined UTI diagnosis process. This user-friendly web application integrates an advanced ensemble model that combines the strengths of LGBM (Light Gradient Boosting Machine) and MLP (Multi-Layer Perceptron) for accurate UTI predictions. Input urinalysis test results effortlessly and receive instant diagnoses, enhancing clinical decision-making. Downloadable results provide convenient record-keeping. Improve patient care with our intuitive and efficient UTI Diagnosis Applicaton.")






    

# Input options
option_time = [datetime.strptime(f"{hour:02d}:{minute:02d} {ampm}", "%I:%M %p").strftime("%I:%M %p")
               for hour in range(1, 13)
               for minute in range(0, 60, 1)
               for ampm in ["AM", "PM"]]
option_spec = [round(x, 3) for x in np.arange(1.005, 1.030, 0.005)]
option_ph = [round(x, 1) for x in np.arange(5.0, 10.0, 1.0)]
option_age = [round(x, 2) for x in np.arange(0.1, 1.0, 0.1)] + list(np.arange(1, 90, 1))
option_ampm = ["AM", "PM"]
option_gender = ["MALE", "FEMALE"]
option_au = ["NONE SEEN", "RARE", "FEW", "OCCASIONAL", "MODERATE", "LOADED", "PLENTY"]
option_color = [
    "LIGHT YELLOW", "STRAW", "AMBER", "BROWN", "DARK YELLOW", "YELLOW", "REDDISH YELLOW", "REDDISH", "LIGHT RED", "RED",
]
option_transparency = ["CLEAR", "SLIGHTLY HAZY", "HAZY", "CLOUDY", "TURBID"]
option_pg = ["NEGATIVE", "TRACE", "1+", "2+", "3+", "4+"]
option_wbc = ["0-0", "0-1", "0-2", "1-2", "0-3", "0-4", "1-3", "2-3", "1-4", "2-4", "1-5", "2-5", "3-4", "1-6", "3-5", "2-6", "3-6", "2-7", "4-5", "3-7", "4-6", "4-7", "5-6", "5-7", "4-8", "3-10", "5-8", "6-8", "4-10", "5-10", "7-8", "7-9", "7-10", "8-10", "8-11", "6-14", "8-12", "9-11", "9-12", "7-15", "10-12", "9-15", "11-13", "10-15", "11-14", "10-16", "11-15", "12-14", "12-15", "10-18", "13-15", "12-17", "14-16", "15-17", "15-18", "16-18", "15-20", "15-21", "17-20", "15-22", "18-20", "18-21", "18-22", "20-22", "15-28", "18-25", "20-25", "22-24", "23-25", "25-30", "25-32", "28-30", "30-32", "28-35", "30-35", "34-36", "36-38", "35-40", "38-40", "45-50", ">50", "48-55", "50-55", "48-62", "55-58", "70-75", "79-85", "85-87", ">100", "LOADED", "TNTC"]
option_rbc = ["0-0", "0-1", "0-2", "1-2", "0-3", "0-4", "1-3", "2-3", "1-4", "2-4", "1-5", "2-5", "3-4", "1-6", "3-5", "2-6", "3-6", "2-7", "4-5", "3-7", "4-6", "4-7", "5-6", "5-7", "4-8", "3-10", "5-8", "6-8", "4-10", "5-10", "7-8", "7-9", "7-10", "8-10", "8-11", "6-14", "8-12", "9-11", "9-12", "7-15", "10-12", "9-15", "11-13", "10-15", "11-14", "10-16", "11-15", "12-14", "12-15", "10-18", "13-15", "12-17", "14-16", "15-17", "15-18", "16-18", "15-20", "15-21", "17-20", "15-22", "18-20", "18-21", "18-22", "20-22", "15-28", "18-25", "20-25", "22-24", "23-25", "25-30", "25-32", "28-30", "30-32", "28-35", "30-35", "34-36", "36-38", "35-40", "38-40", "45-50", ">50", "48-55", "50-55", "48-62", "55-58", "70-75", "79-85", "85-87", ">100", "LOADED", "TNTC"]


st.subheader("Enter the necessary information of the patient :test_tube:", divider='grey')
st.markdown("**Please fill in all fields.**")





# Create two columns
col1, col2 = st.columns(2)

with col1:
    lName = st.text_input("Last Name:", value=None)
    fName = st.text_input("First Name:", value=None)
    mName = st.text_input("Middle Name:", value=None)
    Birthday = st.date_input("Select your birthday:", value=None, min_value=datetime(1600, 1, 1))
    Age = st.selectbox("Age:", [None] + option_age, format_func=lambda x: '' if x is None else x)
    Gender = st.selectbox("Gender:", [None] + option_gender, format_func=lambda x: '' if x is None else x)

    Color = st.selectbox("Color:", [None] + option_color, format_func=lambda x: '' if x is None else x)
    Transparency = st.selectbox("Transparency:", [None] + option_transparency, format_func=lambda x: '' if x is None else x)
    Glucose = st.selectbox("Glucose:", [None] + option_pg, format_func=lambda x: '' if x is None else x)
    Protein = st.selectbox("Protein:", [None] + option_pg, format_func=lambda x: '' if x is None else x)
    pH = st.selectbox("pH:", [None] + option_ph, format_func=lambda x: f"{x:.1f}" if x is not None else '')
    Specific_Gravity = st.selectbox("Specific Gravity:", [None] + option_spec, format_func=lambda x: f"{x:.3f}" if x is not None else '')


# Place input fields in the second column
with col2:
    date = st.date_input("Date Released:", value=None, min_value=datetime(1900, 1, 1))
    time = st.selectbox("Time Released:", [None] + option_time, format_func=lambda x: '' if x is None else x)
    WBC = st.selectbox("White Blood Cells (WBC):", [None] + option_wbc, format_func=lambda x: '' if x is None else x)
    RBC = st.selectbox("Red Blood Cells (RBC):", [None] + option_rbc, format_func=lambda x: '' if x is None else x)
    Epithelial_Cells = st.selectbox("Epithelial Cells:", [None] + option_au, format_func=lambda x: '' if x is None else x)
    Mucous_Threads = st.selectbox("Mucous Threads:", [None] + option_au, format_func=lambda x: '' if x is None else x)
    Amorphous_Urates = st.selectbox("Amorphous Urates:", [None] + option_au, format_func=lambda x: '' if x is None else x)
    Bacteria = st.selectbox("Bacteria:", [None] + option_au, format_func=lambda x: '' if x is None else x)





diagnose_button = st.button("Diagnose", on_click=set_stage, args=(1,))
# Process the input fields and display the diagnosis based on the button click
if st.session_state.stage > 0:
    # Create a Pandas DataFrame from the input data
    user_input = {
        "Age": Age,
        "Gender": Gender,
        "Color": Color,
        "Transparency": Transparency,
        "Glucose": Glucose,
        "Protein": Protein,
        "pH": pH,
        "Specific Gravity": Specific_Gravity,
        "WBC": str(WBC),
        "RBC": str(RBC),
        "Epithelial Cells": Epithelial_Cells,
        "Mucous Threads": Mucous_Threads,
        "Amorphous Urates": Amorphous_Urates,
        "Bacteria": Bacteria
    }
    

    user_input_df = pd.DataFrame([user_input])

    # Create a new DataFrame for Excel
    excel_output_df = user_input_df.copy()

    # Create a list of missing fields
    missing_fields = [field for field, value in {
        "First Name": fName,
        "Last Name": lName,
        "Middle Name": mName,
        "Birthday": Birthday,
        "Date Released": date,
        "Time Released": time,
        "Age": Age,
        "Gender": Gender,
        "Color": Color,
        "Transparency": Transparency,
        "Glucose": Glucose,
        "Protein": Protein,
        "pH": pH,
        "Specific Gravity": Specific_Gravity,
        "WBC": str(WBC),
        "RBC": str(RBC),
        "Epithelial Cells": Epithelial_Cells,
        "Mucous Threads": Mucous_Threads,
        "Amorphous Urates": Amorphous_Urates,
        "Bacteria": Bacteria
    }.items() if value is None]

    # Check if any of the required fields are empty
    if not all(user_input_df.notnull().values.flatten()) or missing_fields:
        missing_fields_str = ", ".join(missing_fields)
        st.error(f"Please fill in all fields ({missing_fields_str}) before diagnosing.")
        st.stop()



    # Make prediction using MLP and LGBM
    pred_MLP = predict_diagnosis_MLP(user_input_df, scaler, model=MLP)
    pred_LGBM = predict_diagnosis_LGBM(user_input_df, scaler, model=LGBM)

    # Assuming pred_MLP is a PyTorch tensor
    normalized_probs = F.softmax(pred_MLP, dim=1)

    
    numpy_array = normalized_probs.detach().numpy()

    pred_ensemble = ((numpy_array * 0.8) + (pred_LGBM * 0.2))

    
    predicted_class = np.argmax(pred_ensemble, axis=1)

    # Existing logic for diagnosis based on model predictions

    if np.any(predicted_class > 0):
        diagnosis = "POSITIVE"
    else:
        diagnosis = "NEGATIVE"

    # Create a new DataFrame with diagnosis result
    diagnosis_df = pd.DataFrame({"Diagnosis": [diagnosis]})


    # Display the diagnosis with conditional formatting
    st.subheader("Diagnosis :stethoscope:", divider='grey')
    if diagnosis == "NEGATIVE":
        st.success(f"The patient is {diagnosis}.")
    else:
        st.error(f"**The patient is {diagnosis}. Advice to see a doctor for medication.**")






    # Create or load Excel template
    template_path = r"C:\Users\Administrator\Downloads\lab_report_template.xlsx"
    try:
        wb = load_workbook(template_path)
    except FileNotFoundError:
        wb = Workbook()

    # Access the active sheet (assuming only one sheet is there)
    sheet = wb.active


    if st.button("Save to Excel"):
            
            # Write user inputs to specific cells in the Excel template
            sheet["C8"] = f"{lName}, {fName} {mName}".upper()
            sheet["C9"] = f"{Age} {'MONTHS OLD' if Age < 1 else 'YEARS OLD'}"
            sheet["E9"] = Gender
            sheet["C10"] = Birthday.strftime("%b-%d-%y").upper()
            sheet["I8"] = date.strftime("%b-%d-%y").upper()  
            sheet["J8"] = time
            sheet["F19"] = Color
            sheet["F20"] = Transparency
            sheet["F23"] = Glucose
            sheet["F24"] = Protein
            sheet["F25"] = pH
            sheet["F26"] = Specific_Gravity
            sheet["F29"] = WBC
            sheet["F30"] = RBC
            sheet["F31"] = Epithelial_Cells
            sheet["F32"] = Mucous_Threads
            sheet["F33"] = Amorphous_Urates
            sheet["F34"] = Bacteria
            sheet["D37"] = f"{diagnosis} {'for UTI'}"


            wb.save("lab_report_generated.xlsx")
            
            st.success("Laboratory Report Excel file saved successfully!")

            # Move to the next stage
            set_stage(2)
